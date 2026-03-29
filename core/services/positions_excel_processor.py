from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

MISSING = {"", "nan", "none", "null", "nat", "n/a", "#n/a", "-"}


def _clean_token(x: object) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() in MISSING else s


def _num(x: object) -> float:
    if x is None:
        return 0.0
    s = str(x).strip()
    if s in {"-", ""}:
        return 0.0
    s = s.replace("\xa0", "").replace(" ", "")
    try:
        if "," in s and "." in s:
            return float(s.replace(",", ""))
        if "," in s and "." not in s:
            return float(s.replace(".", "").replace(",", "."))
        return float(s)
    except Exception:
        cleaned = re.sub(r"[^0-9.,-]", "", s).replace(".", "").replace(",", ".")
        try:
            return float(cleaned) if cleaned not in ("", ".", "-", "--") else 0.0
        except Exception:
            return 0.0


def _detect_header_row(df_raw: pd.DataFrame) -> int | None:
    candidates = ["Account Number", "Name", "Quantity", "Last ($)", "Market Value ($)", "As of"]
    max_scan = min(300, len(df_raw))
    for i in range(max_scan):
        row = [str(v) for v in df_raw.iloc[i].tolist()]
        hits = sum(any(c.lower() in str(cell).lower() for cell in row) for c in candidates)
        if hits >= 4:
            return i
    return None


def _slice_asset_table(df: pd.DataFrame) -> pd.DataFrame:
    key_cols: list[str] = []
    for cand in ["Name", "Description", "Security Description", "Security Name"]:
        if cand in df.columns:
            key_cols.append(cand)
            break
    for cand in ["Quantity", "Qty"]:
        if cand in df.columns:
            key_cols.append(cand)
            break
    for cand in ["Last ($)", "Price", "Unit Price", "Local Price"]:
        if cand in df.columns:
            key_cols.append(cand)
            break
    for cand in ["Market Value ($)", "Value"]:
        if cand in df.columns:
            key_cols.append(cand)
            break
    if len(key_cols) < 3:
        return df.copy()

    empty_run = 0
    end_idx = len(df)
    for i in range(len(df)):
        row = df.iloc[i]
        empties = sum(1 for c in key_cols if _clean_token(row.get(c, "")) == "")
        if empties == len(key_cols):
            empty_run += 1
            if empty_run >= 10:
                end_idx = i - 9
                break
        else:
            empty_run = 0

    core = df.iloc[:end_idx].copy()

    name_col = key_cols[0]
    total_regex = r"(?i)\b(total|subtotal|grand total|total holdings|total portfolio|account total|consolidated total|portfolio total)\b"
    mask_total_name = core[name_col].astype(str).str.contains(total_regex, regex=True, na=False)
    footer_regex = r"(?i)\b(page|disclaimer|holdings for|as of|account summary)\b"
    mask_footer = core[name_col].astype(str).str.contains(footer_regex, regex=True, na=False)

    acct_col = None
    for cand in ["Account Number", "Account", "Acct", "Account #", "AccountNumber"]:
        if cand in core.columns:
            acct_col = cand
            break
    mask_total_acct = False
    if acct_col:
        mask_total_acct = core[acct_col].astype(str).str.strip().str.match(r"(?i)total$")

    mask_all_empty = core[key_cols].applymap(lambda x: _clean_token(x) == "").all(axis=1)

    qty_col = "Quantity" if "Quantity" in core.columns else ("Qty" if "Qty" in core.columns else None)
    mv_col = "Market Value ($)" if "Market Value ($)" in core.columns else ("Value" if "Value" in core.columns else None)
    mask_qm_empty = False
    if qty_col and mv_col:
        mask_qm_empty = (
            core[qty_col].astype(str).str.strip().isin({"", "-"})
            & core[mv_col].astype(str).str.strip().isin({"", "-"})
        )

    mask_drop = mask_total_name | mask_total_acct | mask_footer | mask_all_empty | mask_qm_empty
    return core[~mask_drop].copy()


def _load_depara(depara_path: Path) -> dict[str, str]:
    x = pd.read_excel(depara_path, dtype=str)
    x.columns = [str(c).strip() for c in x.columns]
    c_conta = next((c for c in x.columns if "conta" in c.lower()), None)
    c_wallet = next((c for c in x.columns if "wallet" in c.lower()), None)
    if not c_conta or not c_wallet:
        return {}
    return dict(
        zip(
            x[c_conta].astype(str).str.strip(),
            x[c_wallet].astype(str).str.strip(),
        )
    )


def _consolidate_cash(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    try:
        dref = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors="coerce").max()
    except Exception:
        dref = None

    if pd.notna(dref):
        df["Data"] = df["Data"].fillna(dref.strftime("%d/%m/%Y"))
        df.loc[df["Data"].astype(str).str.strip().eq("nan"), "Data"] = dref.strftime("%d/%m/%Y")

    df_cash = df[df["Caixa"].astype(str).str.upper().eq("SIM")].copy()
    df_pos = df[~df["Caixa"].astype(str).str.upper().eq("SIM")].copy()
    if df_cash.empty:
        return df.copy()

    for c in ["Quant", "SaldoBruto", "Original Value"]:
        df_cash[c] = pd.to_numeric(df_cash[c], errors="coerce").fillna(0.0)

    grp = (
        df_cash.groupby(["Carteira", "Data", "Moeda"], dropna=False)
        .agg(
            Quant=("Quant", "sum"),
            SaldoBruto=("SaldoBruto", "sum"),
            OriginalValue=("Original Value", "sum"),
        )
        .reset_index()
    )
    grp["PU"] = grp.apply(lambda r: (r["SaldoBruto"] / r["Quant"]) if r["Quant"] not in (0, None) else 1.0, axis=1)
    grp["Ativo"] = "USD"
    grp["Caixa"] = "Não"
    grp.rename(columns={"OriginalValue": "Original Value"}, inplace=True)
    grp = grp[["Data", "Carteira", "Ativo", "Quant", "PU", "SaldoBruto", "Original Value", "Caixa", "Moeda"]]

    out = pd.concat([df_pos[grp.columns], grp], ignore_index=True)

    def _date_key(s: object) -> datetime:
        try:
            return datetime.strptime(str(s), "%d/%m/%Y")
        except Exception:
            return datetime.min

    return out.sort_values(
        by=["Data", "Carteira", "Ativo"],
        key=lambda col: col.map(_date_key) if col.name == "Data" else col,
        ignore_index=True,
    )


def process_positions_excel(holdings_path: Path, depara_path: Path, output_dir: Path) -> Path:
    mp = _load_depara(depara_path)

    xls = pd.ExcelFile(holdings_path)
    df_raw = pd.read_excel(holdings_path, sheet_name=xls.sheet_names[0], header=None, dtype=str)
    hdr = _detect_header_row(df_raw)
    if hdr is None:
        raise RuntimeError("Cabecalho nao encontrado.")
    df = pd.read_excel(holdings_path, sheet_name=xls.sheet_names[0], header=hdr, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df_assets = _slice_asset_table(df)

    _data = pd.to_datetime(df_assets.get("As of"), errors="coerce")
    data_ref = _data.max()
    data_str = _data.dt.strftime("%d/%m/%Y")
    if pd.notna(data_ref):
        data_str = data_str.fillna(data_ref.strftime("%d/%m/%Y"))

    acct_col = "Account Number" if "Account Number" in df_assets.columns else "Account"
    carteira = df_assets.get(acct_col, "").astype(str).str.strip()
    if acct_col in df_assets.columns:
        carteira = carteira.map(lambda x: mp.get(x, x))

    def join_ativo(row: pd.Series) -> str:
        return " - ".join(
            [
                x
                for x in [
                    _clean_token(row.get("Name"))
                    or _clean_token(row.get("Description"))
                    or _clean_token(row.get("Security Description")),
                    _clean_token(row.get("Symbol")),
                    _clean_token(row.get("CUSIP")),
                    _clean_token(row.get("ISIN")),
                    _clean_token(row.get("Maturity Date")),
                ]
                if x
            ]
        )

    ativo = df_assets.apply(join_ativo, axis=1)

    qty = df_assets.get("Quantity", df_assets.get("Qty", 0)).apply(_num)
    last = df_assets.get("Last ($)", df_assets.get("Price", 0)).apply(_num)
    mkt = df_assets.get("Market Value ($)", df_assets.get("Value", 0)).apply(_num)
    pu = [(m / q) if (q not in (0, None) and l == 0.0) else l for q, l, m in zip(qty, last, mkt)]
    saldo = [m if m != 0 else (q * p) for m, q, p in zip(mkt, qty, pu)]

    cash_types = {"Cash, MMF and BDP", "Savings & Time Deposits", "Cash"}
    non_cash_exceptions = {"ETFs / CEFs", "Government Securities"}
    is_cash = (
        df_assets.get("Name", "").astype(str).str.strip().str.upper().eq("US DOLLAR")
        | (
            df_assets.get("Product Type", "").isin(list(cash_types))
            & ~df_assets.get("Product Type", "").isin(list(non_cash_exceptions))
        )
    )

    df_norm = pd.DataFrame(
        {
            "Data": data_str,
            "Carteira": carteira,
            "Ativo": ativo,
            "Quant": qty,
            "PU": pu,
            "SaldoBruto": saldo,
            "Original Value": mkt,
            "Caixa": ["Sim" if f else "Não" for f in is_cash],
            "Moeda": "USD",
        }
    )
    df_norm = df_norm[df_norm["Carteira"].astype(str).str.strip().str.match(r"(?i)total$") == False]
    df_final = _consolidate_cash(df_norm)

    ts = datetime.now().strftime("%d-%m-%Y--%H-%M-%S")
    out_xlsx = output_dir / f"positions_processado_V6-2-{ts}.xlsx"

    base_cols = ["Data", "Carteira", "Ativo", "Quant", "PU", "SaldoBruto", "Original Value", "Caixa", "Moeda"]
    col_diff_1 = "Dif (QTD*PU) - Original Value)"
    col_diff_2 = "Saldo Bruto - Original Value"
    for c in [*base_cols, col_diff_1, col_diff_2]:
        if c not in df_final.columns:
            df_final[c] = None if c in ("Quant", "PU", "SaldoBruto", "Original Value", col_diff_1, col_diff_2) else ""
    df_final = df_final[base_cols + [col_diff_1, col_diff_2]]

    mask_usd = df_final["Ativo"].astype(str).str.upper().eq("USD")
    df_final.loc[mask_usd, "Quant"] = df_final.loc[mask_usd, "SaldoBruto"]
    df_final.loc[mask_usd, "PU"] = 1.0

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Processado")
        ws = writer.sheets["Processado"]
        header_map = {cell.value: j for j, cell in enumerate(ws[1], start=1)}

        col_q = header_map.get("Quant")
        col_pu = header_map.get("PU")
        col_sb = header_map.get("SaldoBruto")
        col_ov = header_map.get("Original Value")
        col_dif = header_map.get(col_diff_1)
        col_sbm = header_map.get(col_diff_2)

        max_row = ws.max_row
        if max_row >= 2 and all([col_q, col_pu, col_sb, col_ov]):
            for r in range(2, max_row + 1):
                if col_dif:
                    ws.cell(row=r, column=col_dif).value = (
                        f"=({get_column_letter(col_q)}{r}*{get_column_letter(col_pu)}{r})-{get_column_letter(col_ov)}{r}"
                    )
                if col_sbm:
                    ws.cell(row=r, column=col_sbm).value = (
                        f"={get_column_letter(col_sb)}{r}-{get_column_letter(col_ov)}{r}"
                    )

            grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            for col in (col_ov, col_dif, col_sbm):
                if not col:
                    continue
                ws.cell(row=1, column=col).fill = grey
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col).fill = grey

            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            for c in (col_dif, col_sbm):
                if c:
                    letter = get_column_letter(c)
                    rng = f"{letter}2:{letter}{max_row}"
                    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["5"], fill=red_fill, stopIfTrue=True))
                    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["-5"], fill=red_fill, stopIfTrue=True))

    return out_xlsx
